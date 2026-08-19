"""Ingest BFS Strukturerhebung average rent by canton and room count.

Source: BFS asset 36398436, "Durchschnittlicher Mietpreis nach Zimmerzahl
und Kanton" (T 09.03.03.01). One sheet per year (2010-2024), 26 canton rows
in official BFS canton-number order (row 8 = kt_id 1 ... row 33 = kt_id 26),
row 7 = national total (skipped, not tied to a canton).

Value columns per sheet (0-indexed): Totale=1, room1=3, room2=5, room3=7,
room4=9, room5=11, room6plus=13. Interspersed odd columns are confidence
intervals and are not loaded. 'X' marks a suppressed value (<50 observations
extrapolated) and is loaded as NULL.
"""
from datetime import datetime, timezone

import openpyxl

from ..db import get_connection, init_schema, log_ingest
from ..http import fetch

SOURCE = "rent_bfs"
URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/36398436/master"
ROOM_COLUMNS = {
    "Totale": 1,
    "1": 3,
    "2": 5,
    "3": 7,
    "4": 9,
    "5": 11,
    "6+": 13,
}
FIRST_CANTON_ROW = 8  # kt_id = row - 7
LAST_CANTON_ROW = 33


def _parse_value(v):
    if v is None or isinstance(v, str):
        return None
    return float(v)


def run() -> int:
    path = fetch(URL, SOURCE, filename="mietpreis_kanton.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)

    rows_to_load = []
    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        year = int(sheet_name)
        ws = wb[sheet_name]
        for row_idx in range(FIRST_CANTON_ROW, LAST_CANTON_ROW + 1):
            kt_id = row_idx - (FIRST_CANTON_ROW - 1)
            row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
            for room_cat, col_idx in ROOM_COLUMNS.items():
                value = _parse_value(row[col_idx])
                rows_to_load.append((kt_id, year, room_cat, value, SOURCE, datetime.now(timezone.utc)))

    con = get_connection()
    init_schema(con)
    con.executemany(
        """
        INSERT INTO fact_rent (kt_id, year, room_count_cat, avg_rent_chf, source_dataset, loaded_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT (kt_id, year, room_count_cat) DO UPDATE SET
            avg_rent_chf = excluded.avg_rent_chf,
            source_dataset = excluded.source_dataset,
            loaded_at = excluded.loaded_at
        """,
        rows_to_load,
    )
    log_ingest(con, SOURCE, URL, len(rows_to_load), "ok")
    con.close()
    return len(rows_to_load)


if __name__ == "__main__":
    n = run()
    print(f"Loaded {n} fact_rent rows")
